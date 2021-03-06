import requests
import json
import re
import os
import requests
import openpyxl
import data_vulnerability_classification


### CVE Search

def get_ms_cves_for_date_range(fromPublishedDate, toPublishedDate):
    # Interface for service https://portal.msrc.microsoft.com/en-us/security-guidance
    # fromPublishedDate = "09/09/2020"
    # toPublishedDate = "10/15/2020"

    headers = {
        'authority': 'portal.msrc.microsoft.com',
        'accept': 'application/json, text/plain, */*',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.111 Safari/537.36',
        'content-type': 'application/json;charset=UTF-8',
        'origin': 'https://portal.msrc.microsoft.com',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-mode': 'cors',
        'sec-fetch-dest': 'empty',
        'referer': 'https://portal.msrc.microsoft.com/en-us/security-guidance',
        'accept-language': 'en-US,en;q=0.9,ru;q=0.8',
    }

    data = {"familyIds":[],"productIds":[],"severityIds":[],"impactIds":[],"pageNumber":1,"pageSize":20,
            "includeCveNumber":True,"includeSeverity":False,"includeImpact":False,
            "orderBy":"publishedDate","orderByMonthly":"releaseDate","isDescending":True,"isDescendingMonthly":True,
            "queryText":"","isSearch":False,"filterText":"",
            "fromPublishedDate":fromPublishedDate,"toPublishedDate":toPublishedDate} #Month/Date/Year

    response = requests.post('https://portal.msrc.microsoft.com/api/security-guidance/en-us/excel', headers=headers, data=json.dumps(data))
    f = open("data/ms_search/temp.xlsx", "wb")
    for chunk in response.iter_content(chunk_size=1024):
        if chunk:  # filter out keep-alive new chunks
            f.write(chunk)
    f.close()

    wb = openpyxl.load_workbook(filename="data/ms_search/temp.xlsx")
    ws = wb.worksheets[0]
    row_n = 1
    first_col_value = ""
    cve_ids = set()
    while first_col_value != None:
        first_col_value = ws.cell(column=1, row=row_n).value
        vuln_id = ws.cell(column=7, row=row_n).value
        if vuln_id != None:
            if "CVE-" in vuln_id:
                cve_ids.add(vuln_id)
        row_n += 1
    return(cve_ids)

### CVE Data
def get_ms_cve_data_from_ms_site(cve_id):
    # https://portal.msrc.microsoft.com/en-US/security-guidance/advisory/CVE-2020-1003
    # cve_id = "CVE-2020-1003"
    ms_cve_data = dict()
    try:
        r = requests.get("https://portal.msrc.microsoft.com/api/security-guidance/en-US/CVE/" + cve_id)
        ms_cve_data = r.json()
        ms_cve_data['error'] = False
        ms_cve_data['status'] = "CVE ID was found on microsoft.com portal"
        ms_cve_data['not_found_error'] = False
    except:
        ms_cve_data['error'] = True
        ms_cve_data['status'] = "CVE ID is NOT found on microsoft.com portal"
        ms_cve_data['not_found_error'] = True
    return(ms_cve_data)


def download_ms_cve_data_raw(cve_id, rewrite_flag = True):
    file_path = "data/ms_cve/" + cve_id + ".json"
    if not rewrite_flag:
        if not os.path.exists(file_path):
            # print(cve_id)
            cve_data = get_ms_cve_data_from_ms_site(cve_id)
            f = open(file_path, "w")
            f.write(json.dumps(cve_data))
            f.close()
    else:
        # print(cve_id)
        cve_data = get_ms_cve_data_from_ms_site(cve_id)
        f = open(file_path, "w")
        f.write(json.dumps(cve_data))
        f.close()


def get_ms_cve_data_raw(cve_id):
    f = open("data/ms_cve/" + cve_id + ".json", "r")
    ms_cve_data = json.loads(f.read())
    f.close()
    return(ms_cve_data)


def get_ms_cve_data(cve_id, rewrite_flag):
    download_ms_cve_data_raw(cve_id, rewrite_flag)
    ms_cve_data = get_ms_cve_data_raw(cve_id)
    if ms_cve_data['not_found_error'] == False:
        ms_cve_data = add_cve_product_and_type_tags(ms_cve_data)
        ms_cve_data = heuristic_change_product_vuln_type(ms_cve_data)
        ms_cve_data = add_ms_cve_severity(ms_cve_data)
    return(ms_cve_data)


def add_cve_product_and_type_tags(ms_cve_data):
    for pattern in data_vulnerability_classification.vulnerability_type_detection_patterns:
        if pattern in ms_cve_data['cveTitle']:
            ms_cve_data['vuln_type'] = data_vulnerability_classification.vulnerability_type_detection_patterns[pattern]
            ms_cve_data['vuln_product'] = re.sub( "[ \t]*" + pattern + ".*$", "", ms_cve_data['cveTitle'])
            #print(ms_cve_data[cve_id]['vul_product'] + " - " + ms_cve_data[cve_id]['cveTitle'] )
    if (not 'vuln_type' in ms_cve_data) or (not 'vuln_product' in ms_cve_data):
        print("Error in add_cve_product_and_type_tags for " + ms_cve_data['cveNumber'])
    return(ms_cve_data)


def add_ms_cve_severity(ms_cve_data):
    def MAX(sets):
        return (max(sets))

    severities = set()
    severity_numbers = set()
    severity_dict = {"critical": 4, "important": 3, "moderate": 2, "low": 1}
    result_severity = ""

    for block in ms_cve_data['affectedProducts']:
        severities.add(block['severity'].lower())
    for severity in severities:
        for severity_val in severity_dict:
            if severity == severity_val:
                severity_numbers.add(severity_dict[severity_val])
    max_severity_number = MAX(severity_numbers)
    for severity_val in severity_dict:
        if severity_dict[severity_val] == max_severity_number:
            result_severity = severity_val
    ms_cve_data['severity'] = result_severity
    return(ms_cve_data)


# Heuristics
def heuristic_change_product_vuln_type(ms_cve_data):
    if 'vuln_product' in ms_cve_data:
        if ms_cve_data['vuln_product'] == "Microsoft Office SharePoint":
            ms_cve_data['vuln_product'] = "Microsoft SharePoint"
        if ms_cve_data['vuln_product'] == "Microsoft SharePoint Server":
            ms_cve_data['vuln_product'] = "Microsoft SharePoint"
        if ms_cve_data['vuln_product'] == "Windows VBScript Engine":
            ms_cve_data['vuln_product'] = "VBScript"
        if ms_cve_data['vuln_product'] == "Windows Defender Antimalware Platform Hard Link":
            ms_cve_data['vuln_product'] = "Microsoft Defender"
        if ms_cve_data['vuln_product'] == "Win32k":
            ms_cve_data['vuln_product'] = "Windows Kernel"
        if ms_cve_data['vuln_product'] == "SharePoint":
            ms_cve_data['vuln_product'] = "Microsoft SharePoint"
        if ms_cve_data['vuln_product'] == "Scripting Engine" and \
                "Internet Explorer" in ms_cve_data['description']:
            ms_cve_data['vuln_product'] = "Internet Explorer"
        if ms_cve_data['vuln_product'] == "Scripting Engine" and \
                "ChakraCore scripting engine" in ms_cve_data['description']:
            ms_cve_data['vuln_product'] = "Chakra Scripting Engine"
    if 'vuln_product' in ms_cve_data:
        if ms_cve_data['vuln_type'] == "Memory Corruption" and \
                re.findall("[Rr]emote code execution", ms_cve_data['description']):
            ms_cve_data['vuln_type'] = "Remote Code Execution"
    return(ms_cve_data)

